import openpyxl

# wb = openpyxl.load_workbook('store.xlsx')
# sheet = wb.active

# sheet['d2'] = 400
# new_product = (11, 'tablet', 12, 600, 12*600)
# sheet.append(new_product)
#
# for c, d, e in sheet['c2:e12']:
#     e.value = c.value * d.value
#
# wb.save('store.xlsx')

# wb = openpyxl.Workbook()
#
# sheet = wb.active
#
# sales = {2017:700000, 2018:800000, 2019:900000}
#
# sheet['A1'] = 'Year'
# sheet['B1'] = 'Sales'
#
# for k,v in sales.items():
#     sheet.append((k,v))
#
# wb.save('sales.xlsx')

# wb = openpyxl.load_workbook('store.xlsx')
# sheet = wb.active

# sheet['e3'] = '=c3*d3'
# for c, d, e in sheet['c2:e12']:
#     e.value = f'={c.coordinate}*{d.coordinate}'

# sheet = wb['Products']
# print(dir(sheet))
# sheet.title = 'Products for sale'
# print(sheet.title)
# print(sheet.sheet_format)
# print(sheet.sheet_properties)

# wb.create_sheet('Test 1', 0)
# sheet1 = wb['Test 1']
# wb.remove(sheet1)

# source = wb['Test 1']
# destination = wb.copy_worksheet(sheet)

# wb.save('store.xlsx')

from openpyxl.styles import *
from copy import copy

wb = openpyxl.load_workbook('store.xlsx')
sheet = wb.active

my_cell = sheet['B4']

# print(dir(openpyxl.styles))
# c = Color(indexed=2)
# font1 = Font(name='Tahoma', size=16, color=Color(indexed=2), bold=True, italic=True, strike=False)
# my_cell.font = font1
#
# fill = PatternFill(fill_type='solid', fgColor=Color(indexed=6))
# my_cell.fill = fill
#
# double_border = Side(border_style='double', color=Color(indexed=3))
# thin_border = Side(border_style='thin', color=Color(indexed=4))
# my_cell_border = Border(left=double_border, right=thin_border, top=double_border, bottom=thin_border)
#
# alignment = Alignment(horizontal='right', vertical='center')
# my_cell.alignment = alignment

#           para copiar un estilo, no oolvidar importar copy
# new_cell = sheet['B10']
# new_font = copy(my_cell.font)
# new_font.color = colors=Color(indexed=3)
# new_cell.font = new_font

# wb.save('store.xlsx')

