import openpyxl

wb = openpyxl.load_workbook('store.xlsx', data_only=True) # data_only=True es para mostrar valores y no la formulas

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

sheet1 = wb['Products']
# sheet1 = wb.active

sheet2 = wb['Sales 2018']



c2_cel = sheet1['c2']
b2_cel = sheet1['B2']
#
# print(b2_cel.value, c2_cel.value)
# print(c2_cel.row, c2_cel.column)
#
# print(sheet1['A5'].data_type) # it will return 'n' bcoz is a numeric data type
# print(sheet1['A5'].encoding)
#
# print(sheet1['D4'].parent) # me dice donde estamos trabajando
#
# print(dir(b2_cel)) # muestra todos los comandos posibles
#
# cell_range = sheet1['b2:c11']
# for product, price in cell_range:
#     print(f'Product: {product.value} Price {price.value}')

# Dimensions

print(f'Sheet dimensions {sheet1.dimensions}') # entrega el rnago de celdas que se estan usando

print(sheet1.max_row, sheet1.max_column)

for a, b, c, d, e in sheet1[sheet1.dimensions]:
    print(a.value, b.value, c.value, d.value, e.value,) # muestra formulas

for row in sheet1.rows:
    for cell in row:
        print(f'{cell.value}', end='')
    print('\n')

for row in sheet1.value:
    print(row)