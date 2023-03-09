import openpyxl, json, csv
from openpyxl.styles import *
from copy import copy

# wb = openpyxl.Workbook()
#
# sheet = wb.active
#
# sales = {2017:150000, 2018:180000, 2019:210000, 2020:125000}
#
# sheet['A1'] = 'Year'
# sheet['B1'] = 'Sales'
#
# for k,v in sales.items():
#     sheet.append((k,v))
#
# wb.save('sales.xlsx')

# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = 'COMPANY SALES'
#
# sales = {2017:150000, 2018:180000, 2019:210000, 2020:125000}
#
# sheet['A1'] = 'Year'
# sheet['B1'] = 'Sales'
#
# for k,v in sales.items():
#     sheet.append((k,v))
#
# wb.save('sales.xlsx')

# wb = openpyxl.load_workbook('sales.xlsx')
# sheet1 = wb['COMPANY SALES']
# wb.create_sheet('VAT', 0)
# sheet2 = wb['VAT']
#
# sheet2['A1'] = 'Year'
# sheet2['B1'] = 'Sales'
# sheet2['C1'] = 'VAT'
#
# vat1 = sheet1['a2:c5']
#
# for a, b, c in vat1:
#     c1 = float(b.value) * 0.15
#
#     sheet2.append((a.value, b.value, int(c1)))
#
# wb.save('vat_sales.xlsx')

# wb = openpyxl.load_workbook('sales.xlsx')
# sheet = wb.active
#
# sheet['B6'] = '=sum(b2:b5)'
# sheet['C6'] = 'Total Sales'
#
# wb.save('sales.xlsx')

# wb = openpyxl.load_workbook('sales.xlsx')
# sheet = wb.active
#
# c = Color(indexed=2)
# font1 = Font(name='Tahoma', size=16, color=Color(indexed=2), bold=True, italic=True, strike=False)
#
# sheet['B6'].font = font1
# sheet['C6'].font = font1
#
# wb.save('sales.xlsx')

####################################### version corta!

# def csv2excel(data, result):
#     wb = openpyxl.Workbook()
#     sheet = wb.active
#
#     with open(data, 'r') as f:
#         reader = csv.reader(f)
#         for a in reader:
#             sheet.append(a)
#
#     wb.save(result)
#
# csv2excel('people3.csv', 'teachers.xlsx')

# def csv2excel(data, result, deliter=','):
#     with open(data, 'r') as f:
#         reader = csv.reader(f, delimeter=deliter)
#         for a in reader:
#             # print(x)
#             wb = openpyxl.Workbook()
#             sheet = wb.active
#             sheet['A1'] = 'Person_ID'
#             sheet['B1'] = 'Name'
#             sheet['C1'] = 'First'
#             sheet['D1'] = 'Last'
#             sheet['E1'] = 'Middle'
#             sheet['F1'] = 'Email'
#             sheet['G1'] = 'Phone'
#             sheet['H1'] = 'Fax'
#             sheet['I1'] = 'Title'
#             for x in reader:
#                 sheet.append((x))
#
#     wb.save(result)
#
# csv2excel('people3.csv', 'teachers.xlsx')

# def excel2csv(exl_data, csv_out):
#     wb = openpyxl.load_workbook(exl_data)
#     sheet = wb.active
#     cell_range = sheet['A1:E12']
#     with open(csv_out, 'w', newline='') as f:
#         writer = csv.writer(f)
#         for a,b,c,d,e in cell_range:
#             writer.writerow((a.value,b.value,c.value,d.value,e.value))
#
# excel2csv('books.xlsx', 'booklist.csv')

####################### MODO CORTO

# def excel2csv(exl_data, csv_out):
#     wb = openpyxl.load_workbook(exl_data)
#     sheet = wb.active
#     with open(csv_out, 'w', newline='') as f:
#         writer = csv.writer(f)
#         for row in sheet.values:
#             writer.writerow((row))
#
# excel2csv('books.xlsx', 'booklist.csv')